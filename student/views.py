from django.shortcuts import render, HttpResponse, redirect
from django.db import transaction
from django.http import JsonResponse
from student import models

from openpyxl import load_workbook

from utils.res import ResponseData


def add(request):
    class_queryset = models.Classes.objects.all()
    course_queryset = models.Course.objects.all()
    major_queryset = models.StudentMajor.objects.all()
    tem_dict = {'class_queryset': class_queryset, 'course_queryset': course_queryset,'major_queryset': major_queryset}

    # student = models.Student.objects.get(name="xzh519")
    
    # print(student.student_detail)

    if request.method == 'POST':
        name = request.POST.get('name')
        age = request.POST.get('age', 18)  # 提供默认值
        classes_id = request.POST.get('classes')
        course_id_list = request.POST.getlist('course')
        sex = request.POST.get('sex')
        addr = request.POST.get('province')  # 获取省份
        major = request.POST.get('major')

        # 获取或创建 UserInfo 实例
        user_info, created = models.UserInfo.objects.get_or_create(
            username=name,  # 假设用户名与学生姓名相同
            password="pbkdf2_sha256$390000$ZZFBBAsLEPCVssSfOdoHt3$4amd7bbhQ0wjfhJnbshtPtDuTeV/cxys+wprxJhC6tI=",
            defaults={'role': 'student'}
        )

        try:
            with transaction.atomic():

                majorOB = models.StudentMajor.objects.get(MajorId=major)

                # 创建学生详细信息记录，并与学生记录关联
                detail_obj = models.StudentDetail.objects.create(
                    gender=int(sex),  # 性别转换为整数
                    addr=addr,  # 省份作为地址信息
                    majorId = majorOB,
                      # 关联到刚创建的学生记录
                )
                

                # 创建学生记录
                stu_obj = models.Student.objects.create(
                    name=name, 
                    age=int(age),  # 确保年龄是整数
                    classes_id=classes_id,  # 外键关联班级
                    student_detail=detail_obj,
                    user=user_info,
                )
                stu_obj.save()


                # 处理选课信息
                course_list = []
                for course_id in course_id_list:
                    course_list.append(models.Student2Course(
                        student=stu_obj,
                        course_id=course_id
                    ))
                models.Student2Course.objects.bulk_create(course_list)

        except Exception as e:
            print(e)
            print('服务器错误---student/add!')
            return redirect('base:students')

        return redirect('base:students')

    return render(request, 'student/add.html', tem_dict)

class ResponseData:
    def __init__(self, status=0, message=''):
        self.status = status
        self.message = message

    def get_dict(self):
        return {'status': self.status, 'message': self.message}



def dels(request):
    current_pk = request.POST.get('current_pk')
    if current_pk:
        try:
            student_obj = models.Student.objects.get(pk=current_pk)
            useraccount = student_obj.user
            studentdetail = student_obj.student_detail
            useraccount.delete()
            student_obj.delete()
            studentdetail.delete()
            return JsonResponse({'status': 'success'})
        except models.Student.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found.'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def edit(request, current_pk):

    all_provinces = [
    "北京", "天津", "上海", "重庆",  # 直辖市
    "河北", "山西", "辽宁", "吉林", "黑龙江",  # 东北地区
    "江苏", "浙江", "安徽", "福建", "江西", "山东",  # 华东地区
    "河南", "湖北", "湖南", "广东", "海南",  # 中南地区
    "四川", "贵州", "云南", "陕西", "甘肃", "青海",  # 西部地区
    "内蒙古", "广西", "西藏", "宁夏", "新疆",  # 自治区
    "香港", "澳门"  # 特别行政区
]
   

    # 使用 get() 替代 filter().first()
    stu_obj = models.Student.objects.get(pk=current_pk)
    class_queryset = models.Classes.objects.all()
    course_queryset = models.Course.objects.all()
    # 获取所有专业信息
    all_majors = models.StudentMajor.objects.all()
    

    course_id_list = [course.pk for course in stu_obj.course.all()]

    print(course_id_list)
    

    tem_dict = {
        'stu_obj': stu_obj,
        'class_queryset': class_queryset,
        'course_queryset': course_queryset,
        'course_id_list': course_id_list,
        'all_majors': all_majors,
        'all_provinces': all_provinces,
    }
    if request.method == 'POST':
        name = request.POST.get('name')
        age = request.POST.get('age')
        classes_id = request.POST.get('classes')
        course_id_list = request.POST.getlist('course')
        sex = request.POST.get('sex')
        addr = request.POST.get('province')  # 获取省份
        major_id = request.POST.get('major_id')
        print(sex)
        print(addr)
        
        try:
            with transaction.atomic():
                student = models.Student.objects.select_for_update().get(pk=current_pk)
                student.name = name
                student.age = age
                student.classes_id = classes_id  # 确保这里使用的是正确的字段名
                student.save()

                # 更新学生选课信息
                models.Student2Course.objects.filter(student=student).delete()
                for course_id in course_id_list:
                    models.Student2Course.objects.create(student=student, course_id=course_id)

                # 更新学生详细信息
                student_detail = student.student_detail
                student_detail.gender = sex
                student_detail.majorId = models.StudentMajor.objects.get(pk=major_id)  # 确保这里可以正确获取专业对象
                student_detail.addr = addr
                student_detail.save()

        except Exception as e:
            print(e)
            print('服务器错误---student/edit!')
        finally:
            return redirect('base:students')
    return render(request, 'student/edit.html', tem_dict)


def search(request):
    if request.method == 'POST':
        category = request.POST.get('category')
        key_word = request.POST.get('key_word')
        if category == 'name':
            ret_queryset = models.Student.objects.filter(name__contains=key_word)
        elif category == 'classes':
            ret_queryset = models.Student.objects.filter(classes__name__contains=key_word)
        elif category == 'course':
            ret_queryset = models.Student.objects.filter(course__name=key_word)
        else:
            ret_queryset = None
        return render(request, 'student/search.html', {'ret_queryset': ret_queryset, 'key_word': key_word})


def import_student(request):
    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            return HttpResponse('请先选择文件')
        wb = load_workbook(import_file)
        sheet = wb.worksheets[0]
        gender_dict = {
            '保密': 0,
            '男': 1,
            '女': 2,
        }
        for cells in sheet.iter_rows():
            name = cells[0].value or None
            if name == '姓名':
                continue
            gender = cells[1].value
            gender_val = gender_dict.get(gender, 0)
            age = cells[2].value or None
            addr = cells[3].value or None
            phone = cells[4].value or None
            classes_name = cells[5].value or None
            try:
                with transaction.atomic():
                    class_obj = models.Classes.objects.filter(name=classes_name).first()
                    student_detail_obj = models.StudentDetail.objects.create(gender=gender_val, addr=addr, phone=phone)
                    models.Student.objects.create(name=name, age=age, classes=class_obj,
                                                  student_detail=student_detail_obj)
            except Exception as e:
                print(e)
                print('服务器错误---student/import_student!')
        return redirect('base:students')
